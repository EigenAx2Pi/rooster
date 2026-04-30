"""Generate sample Excel files for testing the web UI.

Produces three files in `samples/`:
  - roster_sample.xlsx        4 months of synthetic booking history (15 employees)
  - holidays_sample.xlsx      a handful of holiday dates in 'DD-MMM' / 'Kochi=holiday' format
  - template_sample.xlsx      blank version of the roster shape, for the optional template path

Re-run with:
    python -m scripts.make_samples
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from app.synth import SynthConfig, generate_history

SAMPLES_DIR = Path(__file__).resolve().parent.parent / "samples"
EMPLOYEE_INFO_HEADERS = [
    "Associate ID", "Associate Name", "Project ID",
    "Project Allocation End Date", "Project Manager ID",
    "Start Time", "End Time", "City", "Facility",
]


def _long_to_wide_workbook(history: pd.DataFrame, *, fill_with_y: bool) -> Workbook:
    """Pivot long-format synthetic history into the wide Excel layout the app expects.

    Layout:
        row 0: employee-info headers in cols 0..8, weekday names in cols 9..N
        row 1: blank in cols 0..8,                 dates (m-d-Y strings) in cols 9..N
        row 2+: one row per employee — info in cols 0..8, 'Y'/blank in cols 9..N
    """
    history = history.copy()
    history["Date"] = pd.to_datetime(history["Date"])

    employees = (
        history[EMPLOYEE_INFO_HEADERS]
        .drop_duplicates(subset=["Associate ID"])
        .sort_values("Associate ID")
        .reset_index(drop=True)
    )
    dates = sorted(history["Date"].unique())

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    for col, header in enumerate(EMPLOYEE_INFO_HEADERS, start=1):
        ws.cell(row=1, column=col).value = header

    for col_offset, d in enumerate(dates):
        col = len(EMPLOYEE_INFO_HEADERS) + 1 + col_offset
        ws.cell(row=1, column=col).value = pd.Timestamp(d).strftime("%A")
        ws.cell(row=2, column=col).value = pd.Timestamp(d).strftime("%m-%d-%Y")

    bookings = history.set_index(["Associate ID", "Date"])["Booked"].to_dict() if fill_with_y else {}

    for emp_idx, emp in employees.iterrows():
        excel_row = 3 + emp_idx
        for col, header in enumerate(EMPLOYEE_INFO_HEADERS, start=1):
            value = emp[header]
            if isinstance(value, pd.Timestamp):
                value = value.strftime("%Y-%m-%d")
            ws.cell(row=excel_row, column=col).value = value
        for col_offset, d in enumerate(dates):
            col = len(EMPLOYEE_INFO_HEADERS) + 1 + col_offset
            booked = bookings.get((emp["Associate ID"], pd.Timestamp(d)), 0)
            if booked == 1:
                ws.cell(row=excel_row, column=col).value = "Y"
    return wb


def _holidays_workbook(year: int) -> Workbook:
    """Holiday file in the format `app.core.get_working_days` expects.

    Two columns the predictor cares about: `Date` (DD-MMM) and `Kochi` (cell == 'holiday').
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Holidays"
    ws.cell(row=1, column=1).value = "Date"
    ws.cell(row=1, column=2).value = "Holiday Name"
    ws.cell(row=1, column=3).value = "Kochi"

    sample_holidays = [
        ("26-Jan", "Republic Day"),
        ("14-Feb", "Local Festival"),
        ("18-Mar", "Holi"),
        ("01-Apr", "Vishu"),
        ("01-May", "Labour Day"),
    ]
    for row_idx, (d, name) in enumerate(sample_holidays, start=2):
        ws.cell(row=row_idx, column=1).value = d
        ws.cell(row=row_idx, column=2).value = name
        ws.cell(row=row_idx, column=3).value = "holiday"
    return wb


def main() -> int:
    SAMPLES_DIR.mkdir(exist_ok=True)
    cfg = SynthConfig(n_employees=15, months=4, start=date(2025, 1, 1), seed=42)
    history = generate_history(cfg)

    roster_wb = _long_to_wide_workbook(history, fill_with_y=True)
    template_wb = _long_to_wide_workbook(history, fill_with_y=False)
    holidays_wb = _holidays_workbook(year=2025)

    roster_path = SAMPLES_DIR / "roster_sample.xlsx"
    template_path = SAMPLES_DIR / "template_sample.xlsx"
    holidays_path = SAMPLES_DIR / "holidays_sample.xlsx"

    roster_wb.save(roster_path)
    template_wb.save(template_path)
    holidays_wb.save(holidays_path)

    print("Wrote:")
    for p in (roster_path, holidays_path, template_path):
        print(f"  {p.relative_to(SAMPLES_DIR.parent)}  ({p.stat().st_size:,} bytes)")
    print("\nTry it:")
    print("  1. Open http://localhost:8000  (or wherever uvicorn is running)")
    print("  2. Upload roster_sample.xlsx as Historical Roster")
    print("  3. Upload holidays_sample.xlsx as Holiday Calendar")
    print("  4. Optionally upload template_sample.xlsx as Blank Template")
    print("  5. Set month=5, year=2025 (held-out month for this synthetic data)")
    print("  6. Click Generate Predictions, download the result.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
